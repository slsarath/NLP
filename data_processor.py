import pandas as pd
import re
import spacy
from langdetect import detect
import logging
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from config.config import LEGAL_PHRASES

class DataProcessor:
    def __init__(self):
        self.nlp = spacy.load('en_core_web_sm', disable=['ner', 'textcat'])
        self.df = None
        self.filtered_df = None
        self.legal_phrases = LEGAL_PHRASES

    def load_data(self, file_path):
        logging.info(f"Loading data from {file_path}")
        try:
            normalized_path = os.path.normpath(file_path)
            self.df = pd.read_excel(normalized_path)
        except Exception as e:
            logging.error(f"Error loading data: {e}")
            raise

    def filter_data(self, creation_date_from, creation_date_to):
        logging.info(f"Filtering data based on creation dates from {creation_date_from} to {creation_date_to} and name")
        self.filtered_df = self.df[
            (self.df['Creation Date'] >= creation_date_from) & 
            (self.df['Creation Date'] <= creation_date_to) & 
            (~self.df['Name'].str.contains("Fraud Cat3"))
        ]

    @staticmethod
    def preprocess_text(text):
        text = str(text).lower()
        text = re.sub(r'[\[\]\\/]', ' ', text)
        text = re.sub(r'\-', '', text)
        text = re.sub(r'\n', '. ', text)
        text = re.sub(r'[\-\;%()|+&=*%!?:#$@\[\]/]', ' ', text)
        return text

    def clean_data(self):
        logging.info("Cleaning data")
        self.filtered_df['Description'] = self.filtered_df['Description'].astype(str)
        self.filtered_df['Risk Event Impact'] = self.filtered_df['Risk Event Impact'].astype(str)
        self.filtered_df['Risk Event Root Cause'] = self.filtered_df['Risk Event Root Cause'].astype(str)

        self.filtered_df['Combined_RE_Narrative'] = self.filtered_df[
            'Description'] + '. ' + self.filtered_df['Risk Event Impact'] + '. ' + self.filtered_df['Risk Event Root Cause']
        
        self.filtered_df['Cleaned_Desc'] = self.filtered_df['Description'].apply(self.preprocess_text)
        self.filtered_df['Cleaned_Impact'] = self.filtered_df['Risk Event Impact'].apply(self.preprocess_text)
        self.filtered_df['Cleaned_Combined_Narrative'] = self.filtered_df['Combined_RE_Narrative'].apply(self.preprocess_text)

    def extract_legal_phrases(self, text):
        extracted_phrases = [phrase for phrase in self.legal_phrases if phrase in text.lower()]
        return extracted_phrases

    def process_legal_phrases(self):
        logging.info("Extracting legal phrases")
        self.filtered_df['Legal_Phrases'] = self.filtered_df['Cleaned_Combined_Narrative'].apply(self.extract_legal_phrases)

    def aggregate_check(self):
        logging.info("Checking aggregate events")
        aggregate_pattern = r'(\d+ accounts|\d+ transactions|multiple accounts|multiple transactions|Aggregated|Aggregate)'
        self.filtered_df['Aggregate_phrases'] = self.filtered_df['Cleaned_Combined_Narrative'].apply(
            lambda x: ','.join(re.findall(aggregate_pattern, x))
        )

    @staticmethod
    def legal_function_check(check_column, legal_phrases):
        if check_column == 'Yes':
            return "Need to be validated"
        elif check_column == 'No' and len(legal_phrases) != 0:
            return "FAIL"
        else:
            return "PASS"

    def apply_legal_check(self):
        logging.info("Applying legal function check")
        self.filtered_df['Function_Check'] = self.filtered_df.apply(
            lambda row: self.legal_function_check(row['Managed By Legal Function'], row['Legal_Phrases']), axis=1
        )

    @staticmethod
    def aggregate_events_check(check_column, aggregate_phrase):
        if check_column == 'Yes':
            return "PASS"
        elif check_column == 'No' and aggregate_phrase.strip() != '':
            return "FAIL"
        else:
            return "PASS"

    def apply_aggregate_check(self):
        logging.info("Applying aggregate events check")
        self.filtered_df['Aggregate_Check'] = self.filtered_df.apply(
            lambda row: self.aggregate_events_check(row['Aggregate Event'], row['Aggregate_phrases']), axis=1
        )

    @staticmethod
    def anticipated_loss_check(check_column, aadl_column):
        if check_column == 'Closed' and aadl_column == 0:
            return "PASS"
        elif check_column == 'Closed' and aadl_column != 0:
            return "FAIL"
        else:
            return "PASS"

    def apply_loss_check(self):
        logging.info("Applying anticipated loss check")
        self.filtered_df['AADL_Check'] = self.filtered_df.apply(
            lambda row: self.anticipated_loss_check(row['Workflow Status'], row['Anticipated Additional Direct (GBP)']), axis=1
        )

    def remove_acronyms(self, text):
        acronym_pattern = r'\b[A-Z]{2,5}\b'
        cleaned_text = re.sub(acronym_pattern, '', text)
        return cleaned_text

    def is_english(self, text):
        cleaned_text = self.remove_acronyms(text)
        try:
            language = detect(cleaned_text)
            return 'Yes' if language == 'en' else 'No'
        except Exception as e:
            logging.error(f"Error detecting language: {e}")
            return 'No'

    def apply_language_check(self):
        logging.info("Checking if RE_Narrative is in English")
        self.filtered_df['Is_RE_Narrative_in_English'] = self.filtered_df['Combined_RE_Narrative'].apply(self.is_english)

    def save_output(self, output_path):
        logging.info(f"Saving output to {output_path}")
        try:
            self.filtered_df.to_excel(output_path, index=False)
            self.apply_color_coding(output_path)
            self.create_pivot_tables(output_path)
        except Exception as e:
            logging.error(f"Error saving output: {e}")
            raise

    def apply_color_coding(self, file_path):
        logging.info(f"Applying color coding to {file_path}")
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            amber_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value == "PASS":
                        cell.fill = green_fill
                    elif cell.value == "FAIL":
                        cell.fill = amber_fill

            wb.save(file_path)
        except Exception as e:
            logging.error(f"Error applying color coding: {e}")
            raise

    def create_pivot_tables(self, file_path):
        logging.info(f"Creating pivot tables in {file_path}")
        try:
            wb = load_workbook(file_path)
            ws = wb.create_sheet(title="Summary")

            # Pivot table for Managed by Legal Function Check
            self.create_pivot_table(
                wb, ws, "Managed by Legal Function Check", "Data!A1:AX100", "Data!A1:A100", "Data!AY1:AY100", 1, 1
            )
            
            # Pivot table for Aggregate Check
            self.create_pivot_table(
                wb, ws, "Aggregate Check", "Data!A1:AX100", "Data!A1:A100", "Data!AZ1:AZ100", 1, 15
            )

            # Pivot table for AADL Check
            self.create_pivot_table(
                wb, ws, "AADL Check", "Data!A1:AX100", "Data!A1:A100", "Data!BA1:BA100", 1, 29
            )

            wb.save(file_path)
        except Exception as e:
            logging.error(f"Error creating pivot tables: {e}")
            raise

    def create_pivot_table(self, wb, ws, title, data_range, row_range, col_range, start_row, start_col):
        # Note: Openpyxl does not directly support pivot tables, 
        # so this is a placeholder for the logic.
        # You might need to use a library like xlwings or manipulate
        # the file with Excel VBA macros.
        pass